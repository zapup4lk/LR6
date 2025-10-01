import openpyxl
import os


'''
Первое изменение для ЛР6
'''

def process_time_files(template_path, output_folder):
    os.makedirs(output_folder, exist_ok=True)

    base_wb = openpyxl.load_workbook(template_path)

    n = 2
    for day in range(1, 32):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for sheet in base_wb.worksheets:
            new_sheet = wb.create_sheet(sheet.title)
            for row in sheet.iter_rows():
                for cell in row:
                    new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)


        for sheet in wb.worksheets:
            for row in sheet.iter_rows(min_col=2, max_col=2):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.value += 86400 * day


        output_path = os.path.join(output_folder, f'658 {n}.wln')
        n += 1


        with open(output_path, 'w', encoding='utf-8') as f:
            for sheet in wb.worksheets:
                for row in sheet.iter_rows():
                    # Формируем строку для записи в файл
                    row_data = [str(cell.value) if cell.value is not None else '' for cell in row]
                    f.write(';'.join(row_data) + '\n')

        wb.close()

    base_wb.close()


process_time_files("903.xlsx", r"C:\Users\zapup4lk\Desktop\903")